import re
import pandas as pd
import numpy as np
from openpyxl import Workbook
import datetime


def list_for_summary_short(data_list):
    sentence_set = set()
    for data in data_list:
        for s in re.split('。|！|？|；| |>|【|】|，|；|：|―|', data):
            if s not in sentence_set and len(s) != 0:
                sentence_set.add(str(s).replace(" ", "").lower().strip())
    return list(sentence_set)


begin = datetime.datetime.now()
data = pd.read_excel("E:/桌面转移/一些文档/近期工作/比亚迪/初期数据/autohome_bbs_2017-10-01.xls")
comment = np.array(data['thread_context']).tolist()
cars = np.array(data['car_series']).tolist()
cut_sent = list_for_summary_short(comment)

car_words = []
emo_words = []
with open('C:/Users/Administrator/Desktop/all_car.txt', 'r', encoding='utf-8') as f:
    for line in f:
        if len(line) != 0:
            car_words.append(str(line).replace(" ", "").lower().strip())

with open('C:/Users/Administrator/Desktop/emo_words.txt', 'r', encoding='utf-8') as f:
    for line in f:
        if len(line) != 0:
            emo_words.append(str(line).replace(" ", "").lower().strip())

print("共读取了%d个长句" % len(comment))
print("切分后共有%d个不重复的分句" % len(cut_sent))
print("车型表中共%d个词" % len(car_words))
print("情感表中共%d个词" % len(emo_words))
print("---------------------------------------")

result = []
result2 = []
count = 0
count2 = 0

for i in cut_sent:
    res = []
    for j in car_words:
        if str(i).__contains__(j):
            res.append(j)
    count += 1
    if count % 1000 == 0:
        print(count)
    result.append(res)

for i in cut_sent:
    res = []
    for j in emo_words:
        if str(i).replace(" ", "").lower().__contains__(j):
            res.append(j)
    count2 += 1
    if count2 % 1000 == 0:
        print(count2)
    result2.append(res)

print("匹配完成，开始写入文件")

try:
    out_wb = Workbook()
    out_ws = out_wb.create_sheet(title="result")

    out_ws.cell(row=1, column=1).value = '分句'
    out_ws.cell(row=1, column=2).value = '包含车型词'
    out_ws.cell(row=1, column=3).value = '包含情感词'

    row_out = 2
    for i in range(len(cut_sent)):
        out_ws.cell(row=i + row_out, column=1).value = str(cut_sent[i]).strip()
        if len(result[i]) != 0:
            out_ws.cell(row=i + row_out, column=2).value = str(result[i])
        else:
            out_ws.cell(row=i + row_out, column=2).value = ""
        if len(result2[i]) != 0:
            out_ws.cell(row=i + row_out, column=3).value = str(result2[i])
        else:
            out_ws.cell(row=i + row_out, column=3).value = ""

    out_wb.save("C:/Users/Administrator/Desktop/匹配结果_all.xlsx")
except Exception as e:
    print(e)
print("done")

step4 = datetime.datetime.now()
print("总共耗时："+str(step4-begin))
