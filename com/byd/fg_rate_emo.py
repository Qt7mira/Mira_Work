import re
import pandas as pd
import numpy as np
import xlwt
from openpyxl import Workbook


def list_for_summary_short(data_list):
    sentence_set = set()
    for data in data_list:
        for s in re.split('。|！|？|；| |>|【|】|，|；|：', data):
            if s not in sentence_set and len(s) != 0:
                sentence_set.add(s)
                # yield s
    return list(sentence_set)


data = pd.read_excel("C:/Users/Administrator/Desktop/autohome_koubei#.xlsx")
comment = np.array(data['comment_content']).tolist()
cut_sent = list_for_summary_short(comment)

emo_words = []
with open('C:/Users/Administrator/Desktop/emo_words.txt', 'r', encoding='utf-8') as f:
    for line in f:
        if len(line) != 0:
            emo_words.append(line.strip())

print("共读取了%d个长句" % len(comment))
print("切分后共有%d个不重复的分句" % len(cut_sent))
print("情感词表中共%d个词" % len(emo_words))
print("---------------------------------------")

result = []
count = 0

for i in cut_sent:
    res = []
    for j in emo_words:
        if str(i).__contains__(j):
            res.append(j)
    count += 1
    if count % 1000 == 0:
        print(count)
    result.append(res)

print("匹配完成，开始写入文件")

out_wb = Workbook()
out_ws = out_wb.create_sheet(title="result")

out_ws.cell(row=1, column=1).value = '分句'
out_ws.cell(row=1, column=2).value = '包含情感词'

row_out = 2
for i in range(len(cut_sent)):
    out_ws.cell(row=i + row_out, column=1).value = str(cut_sent[i])
    if len(result[i]) != 0:
        out_ws.cell(row=i + row_out, column=2).value = str(result[i])
    else:
        out_ws.cell(row=i + row_out, column=2).value = ""

out_wb.save("C:/Users/Administrator/Desktop/匹配结果_1225.xlsx")
print("done")
